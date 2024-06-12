import json
import psycopg2
from openpyxl import load_workbook

def load_config(file_path):
    with open(file_path, 'r') as file:
        return json.load(file)

def create_table(cursor, column_names):
    try:
        cursor.execute("DROP TABLE IF EXISTS movies;")
        cursor.execute("CREATE TABLE movies (id SERIAL PRIMARY KEY);")
        for column_name in column_names:
            cursor.execute(f"ALTER TABLE movies ADD COLUMN {column_name} VARCHAR;")
    except Exception as e:
        print(f"Error creating table: {e}")

def insert_data(cursor, data):
    try:
        for row in data:
            print(f"Inserting data: {row}")  # 用于调试的打印语句
            cursor.execute("INSERT INTO movies VALUES (DEFAULT, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)", row)
        print("All data inserted successfully!")  # 用于调试的打印语句
    except Exception as e:
        print(f"Error inserting data: {e}")

if __name__ == "__main__":
    config = load_config('../config.json')
    db_config = config['database']

    conn = psycopg2.connect(
        dbname=db_config['dbname'],
        user=db_config['user'],
        password=db_config['password'],
        host=db_config['host']
    )
    cursor = conn.cursor()

    # 从 Excel 文件中读取数据
    wb = load_workbook(filename='data_movie.xlsx')
    ws = wb.active
    column_names = [cell.value for cell in ws[1]]
    data = []
    for row in ws.iter_rows(min_row=2):
        row_data = [cell.value for cell in row]
        movie_name = row_data[column_names.index('movie_name')]
        cover_url = f"/img/cover/{movie_name.split(' ')[0]}.jpg"
        row_data.append(cover_url)  # 在每一行数据中加入 coverUrl
        data.append(row_data)

    # 创建表格并插入数据
    create_table(cursor, column_names + ['coverUrl'])  # 添加 coverUrl 列
    insert_data(cursor, data)

    # 提交并关闭连接
    conn.commit()
    cursor.close()
    conn.close()

    print("Data inserted successfully!")
